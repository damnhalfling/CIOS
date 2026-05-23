"""Skill: spreadsheet — read and edit spreadsheet files (CSV, XLSX, ODS).

Supports:
- Reading cell values by column/row
- Searching for values across sheets
- Updating specific cells
- Listing sheet structure (columns, row count)

No LLM. Pure file manipulation.
"""

import csv
import logging
import os
from dataclasses import dataclass
from pathlib import Path

logger = logging.getLogger(__name__)

# Supported extensions
_SUPPORTED_EXTENSIONS = {".csv", ".xlsx", ".xls", ".ods", ".tsv"}

# Default search directories
_SEARCH_DIRS = [
    "~/Documents",
    "~/Downloads",
    "~/Desktop",
    "~",
]


@dataclass
class CellResult:
    """A single cell match from a spreadsheet search."""

    file: str
    sheet: str
    row: int
    column: str
    value: str


@dataclass
class SheetInfo:
    """Metadata about a spreadsheet."""

    file: str
    sheets: list[str]
    columns: list[str]
    row_count: int


@dataclass
class SpreadsheetResult:
    """Result of a spreadsheet operation."""

    success: bool
    plan_steps: list[str]
    message: str
    data: list[CellResult] | SheetInfo | None = None
    error: str | None = None


def find_spreadsheet(name: str) -> str | None:
    """Find a spreadsheet file by name (fuzzy match).

    Searches common directories for files matching the name.
    Returns the full path or None.
    """
    name_lower = name.lower()
    # Remove extension if user provided one
    name_stem = Path(name_lower).stem if "." in name_lower else name_lower

    candidates: list[tuple[str, int]] = []

    for search_dir in _SEARCH_DIRS:
        expanded = os.path.expanduser(search_dir)
        if not os.path.isdir(expanded):
            continue

        for root, dirs, files in os.walk(expanded):
            # Skip hidden and common noise directories
            dirs[:] = [
                d
                for d in dirs
                if not d.startswith(".") and d not in {"node_modules", "__pycache__", ".venv"}
            ]

            for f in files:
                ext = os.path.splitext(f)[1].lower()
                if ext not in _SUPPORTED_EXTENSIONS:
                    continue

                f_stem = Path(f).stem.lower()

                # Exact match
                if f_stem == name_stem:
                    return os.path.join(root, f)

                # Fuzzy: name is contained in filename
                if name_stem in f_stem or f_stem in name_stem:
                    score = abs(len(f_stem) - len(name_stem))
                    candidates.append((os.path.join(root, f), score))

    if candidates:
        candidates.sort(key=lambda x: x[1])
        return candidates[0][0]

    return None


def read_spreadsheet(file_path: str) -> SpreadsheetResult:
    """Read a spreadsheet and return its structure and content summary."""
    path = Path(file_path)
    if not path.exists():
        return SpreadsheetResult(
            success=False,
            plan_steps=["Localizando arquivo"],
            message=f"Arquivo não encontrado: {path.name}",
            error="file_not_found",
        )

    ext = path.suffix.lower()

    try:
        if ext == ".csv":
            return _read_csv(file_path)
        elif ext == ".tsv":
            return _read_csv(file_path, delimiter="\t")
        elif ext in (".xlsx", ".xls"):
            return _read_xlsx(file_path)
        elif ext == ".ods":
            return _read_ods(file_path)
        else:
            return SpreadsheetResult(
                success=False,
                plan_steps=["Verificando formato"],
                message=f"Formato não suportado: {ext}",
                error="unsupported_format",
            )
    except Exception as e:
        logger.warning("Failed to read spreadsheet %s: %s", file_path, e)
        return SpreadsheetResult(
            success=False,
            plan_steps=["Lendo arquivo"],
            message="Não consegui ler o arquivo.",
            error=str(e),
        )


def search_value(file_path: str, query: str) -> SpreadsheetResult:
    """Search for a value in a spreadsheet.

    Searches all cells for the query string (case-insensitive).
    Returns matching cells with their location.
    """
    path = Path(file_path)
    if not path.exists():
        return SpreadsheetResult(
            success=False,
            plan_steps=["Localizando arquivo"],
            message=f"Arquivo não encontrado: {path.name}",
            error="file_not_found",
        )

    ext = path.suffix.lower()
    query_lower = query.lower()
    results: list[CellResult] = []

    try:
        if ext == ".csv":
            results = _search_csv(file_path, query_lower)
        elif ext == ".tsv":
            results = _search_csv(file_path, query_lower, delimiter="\t")
        elif ext in (".xlsx", ".xls"):
            results = _search_xlsx(file_path, query_lower)
        elif ext == ".ods":
            results = _search_ods(file_path, query_lower)

        if results:
            return SpreadsheetResult(
                success=True,
                plan_steps=["Buscando na planilha", f"Encontrados {len(results)} resultados"],
                message=f"Encontrei {len(results)} resultado(s) para '{query}'.",
                data=results,
            )
        else:
            return SpreadsheetResult(
                success=True,
                plan_steps=["Buscando na planilha"],
                message=f"Nenhum resultado encontrado para '{query}'.",
                data=[],
            )
    except Exception as e:
        logger.warning("Failed to search spreadsheet %s: %s", file_path, e)
        return SpreadsheetResult(
            success=False,
            plan_steps=["Buscando na planilha"],
            message="Não consegui buscar no arquivo.",
            error=str(e),
        )


def update_cell(
    file_path: str,
    search_value_str: str,
    new_value: str,
    column_hint: str | None = None,
) -> SpreadsheetResult:
    """Update a cell value in a spreadsheet.

    Finds the cell containing search_value_str and replaces it with new_value.
    If column_hint is provided, only searches in that column.

    Returns success/failure with details.
    """
    path = Path(file_path)
    if not path.exists():
        return SpreadsheetResult(
            success=False,
            plan_steps=["Localizando arquivo"],
            message=f"Arquivo não encontrado: {path.name}",
            error="file_not_found",
        )

    ext = path.suffix.lower()

    try:
        if ext == ".csv":
            return _update_csv(file_path, search_value_str, new_value, column_hint)
        elif ext == ".tsv":
            return _update_csv(file_path, search_value_str, new_value, column_hint, delimiter="\t")
        elif ext in (".xlsx", ".xls"):
            return _update_xlsx(file_path, search_value_str, new_value, column_hint)
        elif ext == ".ods":
            return SpreadsheetResult(
                success=False,
                plan_steps=["Verificando formato"],
                message="Edição de .ods ainda não suportada. Converta para .xlsx ou .csv.",
                error="unsupported_edit",
            )
        else:
            return SpreadsheetResult(
                success=False,
                plan_steps=["Verificando formato"],
                message=f"Formato não suportado para edição: {ext}",
                error="unsupported_format",
            )
    except Exception as e:
        logger.warning("Failed to update spreadsheet %s: %s", file_path, e)
        return SpreadsheetResult(
            success=False,
            plan_steps=["Atualizando planilha"],
            message="Não consegui atualizar o arquivo.",
            error=str(e),
        )


def get_column_value(file_path: str, row_query: str, column_name: str) -> SpreadsheetResult:
    """Get a specific column value from a row matching a query.

    Useful for: "quanto gastamos no último mês?" → find row with "último mês",
    return value from "custo" column.
    """
    path = Path(file_path)
    if not path.exists():
        return SpreadsheetResult(
            success=False,
            plan_steps=["Localizando arquivo"],
            message=f"Arquivo não encontrado: {path.name}",
            error="file_not_found",
        )

    ext = path.suffix.lower()
    row_query_lower = row_query.lower()
    col_lower = column_name.lower()

    try:
        if ext == ".csv":
            rows, headers = _load_csv_rows(file_path)
        elif ext == ".tsv":
            rows, headers = _load_csv_rows(file_path, delimiter="\t")
        elif ext in (".xlsx", ".xls"):
            rows, headers = _load_xlsx_rows(file_path)
        else:
            return SpreadsheetResult(
                success=False,
                plan_steps=["Verificando formato"],
                message=f"Formato não suportado: {ext}",
                error="unsupported_format",
            )

        # Find column index
        col_idx = None
        for i, h in enumerate(headers):
            if col_lower in h.lower():
                col_idx = i
                break

        if col_idx is None:
            return SpreadsheetResult(
                success=False,
                plan_steps=["Buscando coluna"],
                message=f"Coluna '{column_name}' não encontrada. "
                f"Colunas disponíveis: {', '.join(headers)}",
                error="column_not_found",
            )

        # Find row matching query
        for row_num, row in enumerate(rows, start=2):
            row_text = " ".join(str(cell) for cell in row).lower()
            if row_query_lower in row_text:
                value = row[col_idx] if col_idx < len(row) else ""
                return SpreadsheetResult(
                    success=True,
                    plan_steps=["Buscando na planilha", "Valor encontrado"],
                    message=str(value),
                    data=[
                        CellResult(
                            file=path.name,
                            sheet="Sheet1",
                            row=row_num,
                            column=headers[col_idx],
                            value=str(value),
                        )
                    ],
                )

        return SpreadsheetResult(
            success=True,
            plan_steps=["Buscando na planilha"],
            message=f"Nenhuma linha encontrada com '{row_query}'.",
            data=[],
        )

    except Exception as e:
        logger.warning("Failed to get column value from %s: %s", file_path, e)
        return SpreadsheetResult(
            success=False,
            plan_steps=["Lendo planilha"],
            message="Não consegui ler o valor.",
            error=str(e),
        )


# ═══════════════════════════════════════════════════════════════════════════
#  CSV OPERATIONS
# ═══════════════════════════════════════════════════════════════════════════


def _read_csv(file_path: str, delimiter: str = ",") -> SpreadsheetResult:
    """Read a CSV file and return structure info."""
    with open(file_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f, delimiter=delimiter)
        rows = list(reader)

    if not rows:
        return SpreadsheetResult(
            success=True,
            plan_steps=["Lendo arquivo"],
            message="Arquivo vazio.",
            data=SheetInfo(file=file_path, sheets=["Sheet1"], columns=[], row_count=0),
        )

    headers = rows[0]
    row_count = len(rows) - 1  # exclude header

    return SpreadsheetResult(
        success=True,
        plan_steps=["Lendo arquivo", f"{row_count} linhas, {len(headers)} colunas"],
        message=f"Planilha com {row_count} linhas. Colunas: {', '.join(headers)}",
        data=SheetInfo(file=file_path, sheets=["Sheet1"], columns=headers, row_count=row_count),
    )


def _load_csv_rows(file_path: str, delimiter: str = ",") -> tuple[list[list[str]], list[str]]:
    """Load CSV rows and headers."""
    with open(file_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f, delimiter=delimiter)
        rows = list(reader)

    if not rows:
        return [], []

    headers = rows[0]
    data_rows = rows[1:]
    return data_rows, headers


def _search_csv(file_path: str, query: str, delimiter: str = ",") -> list[CellResult]:
    """Search all cells in a CSV for a query string."""
    results: list[CellResult] = []

    with open(file_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f, delimiter=delimiter)
        rows = list(reader)

    if not rows:
        return results

    headers = rows[0]

    for row_idx, row in enumerate(rows[1:], start=2):
        for col_idx, cell in enumerate(row):
            if query in str(cell).lower():
                col_name = headers[col_idx] if col_idx < len(headers) else f"Col{col_idx}"
                results.append(
                    CellResult(
                        file=Path(file_path).name,
                        sheet="Sheet1",
                        row=row_idx,
                        column=col_name,
                        value=str(cell),
                    )
                )

    return results


def _update_csv(
    file_path: str,
    search_val: str,
    new_val: str,
    column_hint: str | None = None,
    delimiter: str = ",",
) -> SpreadsheetResult:
    """Update a cell in a CSV file."""
    with open(file_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f, delimiter=delimiter)
        rows = list(reader)

    if not rows:
        return SpreadsheetResult(
            success=False,
            plan_steps=["Lendo arquivo"],
            message="Arquivo vazio.",
            error="empty_file",
        )

    headers = rows[0]
    search_lower = search_val.lower()
    updated = False
    updated_location = ""

    # Find column index if hint provided
    target_col = None
    if column_hint:
        col_hint_lower = column_hint.lower()
        for i, h in enumerate(headers):
            if col_hint_lower in h.lower():
                target_col = i
                break

    for row_idx, row in enumerate(rows[1:], start=1):
        for col_idx, cell in enumerate(row):
            # Skip if column hint doesn't match
            if target_col is not None and col_idx != target_col:
                continue

            if search_lower in str(cell).lower():
                rows[row_idx][col_idx] = new_val
                col_name = headers[col_idx] if col_idx < len(headers) else f"Col{col_idx}"
                updated_location = f"linha {row_idx + 1}, coluna '{col_name}'"
                updated = True
                break
        if updated:
            break

    if not updated:
        return SpreadsheetResult(
            success=False,
            plan_steps=["Buscando valor na planilha"],
            message=f"Valor '{search_val}' não encontrado na planilha.",
            error="value_not_found",
        )

    # Write back
    with open(file_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, delimiter=delimiter)
        writer.writerows(rows)

    return SpreadsheetResult(
        success=True,
        plan_steps=["Buscando valor", "Atualizando", "Salvando"],
        message=f"Atualizado: {updated_location} → '{new_val}'.",
    )


# ═══════════════════════════════════════════════════════════════════════════
#  XLSX OPERATIONS (requires openpyxl)
# ═══════════════════════════════════════════════════════════════════════════


def _get_openpyxl():
    """Import openpyxl with graceful degradation."""
    try:
        import openpyxl

        return openpyxl
    except ImportError:
        return None


def _read_xlsx(file_path: str) -> SpreadsheetResult:
    """Read an XLSX file and return structure info."""
    openpyxl = _get_openpyxl()
    if not openpyxl:
        return SpreadsheetResult(
            success=False,
            plan_steps=["Verificando dependências"],
            message="openpyxl não instalado. Instale com: pip install openpyxl",
            error="missing_dependency",
        )

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    ws = wb.active

    headers = []
    row_count = 0

    if ws:
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = [str(cell) if cell else "" for cell in row]
        row_count = ws.max_row - 1 if ws.max_row else 0

    wb.close()

    return SpreadsheetResult(
        success=True,
        plan_steps=["Lendo arquivo", f"{row_count} linhas, {len(headers)} colunas"],
        message=f"Planilha com {row_count} linhas, {len(sheet_names)} aba(s). "
        f"Colunas: {', '.join(headers)}",
        data=SheetInfo(file=file_path, sheets=sheet_names, columns=headers, row_count=row_count),
    )


def _load_xlsx_rows(file_path: str) -> tuple[list[list], list[str]]:
    """Load XLSX rows and headers."""
    openpyxl = _get_openpyxl()
    if not openpyxl:
        return [], []

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    rows = []
    headers = []

    if ws:
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(cell) if cell else "" for cell in row]
            else:
                rows.append([str(cell) if cell is not None else "" for cell in row])

    wb.close()
    return rows, headers


def _search_xlsx(file_path: str, query: str) -> list[CellResult]:
    """Search all cells in an XLSX file."""
    openpyxl = _get_openpyxl()
    if not openpyxl:
        return []

    results: list[CellResult] = []
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = []

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_idx == 1:
                headers = [str(cell) if cell else f"Col{i}" for i, cell in enumerate(row)]
                continue

            for col_idx, cell in enumerate(row):
                if cell is not None and query in str(cell).lower():
                    col_name = headers[col_idx] if col_idx < len(headers) else f"Col{col_idx}"
                    results.append(
                        CellResult(
                            file=Path(file_path).name,
                            sheet=sheet_name,
                            row=row_idx,
                            column=col_name,
                            value=str(cell),
                        )
                    )

    wb.close()
    return results


def _update_xlsx(
    file_path: str,
    search_val: str,
    new_val: str,
    column_hint: str | None = None,
) -> SpreadsheetResult:
    """Update a cell in an XLSX file."""
    openpyxl = _get_openpyxl()
    if not openpyxl:
        return SpreadsheetResult(
            success=False,
            plan_steps=["Verificando dependências"],
            message="openpyxl não instalado. Instale com: pip install openpyxl",
            error="missing_dependency",
        )

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    search_lower = search_val.lower()
    updated = False
    updated_location = ""

    # Get headers
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = [str(cell) if cell else "" for cell in row]

    # Find target column if hint provided
    target_col = None
    if column_hint:
        col_hint_lower = column_hint.lower()
        for i, h in enumerate(headers):
            if col_hint_lower in h.lower():
                target_col = i
                break

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for col_idx, cell in enumerate(row):
            if target_col is not None and col_idx != target_col:
                continue

            cell_val = str(cell.value) if cell.value is not None else ""
            if search_lower in cell_val.lower():
                # Try to preserve type (number vs string)
                try:
                    cell.value = float(new_val) if "." in new_val else int(new_val)
                except ValueError:
                    cell.value = new_val

                col_name = headers[col_idx] if col_idx < len(headers) else f"Col{col_idx}"
                updated_location = f"linha {row_idx}, coluna '{col_name}'"
                updated = True
                break
        if updated:
            break

    if not updated:
        wb.close()
        return SpreadsheetResult(
            success=False,
            plan_steps=["Buscando valor na planilha"],
            message=f"Valor '{search_val}' não encontrado na planilha.",
            error="value_not_found",
        )

    wb.save(file_path)
    wb.close()

    return SpreadsheetResult(
        success=True,
        plan_steps=["Buscando valor", "Atualizando", "Salvando"],
        message=f"Atualizado: {updated_location} → '{new_val}'.",
    )


# ═══════════════════════════════════════════════════════════════════════════
#  ODS OPERATIONS (read-only via odfpy or csv export)
# ═══════════════════════════════════════════════════════════════════════════


def _read_ods(file_path: str) -> SpreadsheetResult:
    """Read an ODS file (basic support)."""
    # Try to use pandas if available, otherwise report unsupported
    try:
        import pandas as pd

        df = pd.read_excel(file_path, engine="odf")
        headers = list(df.columns)
        row_count = len(df)

        return SpreadsheetResult(
            success=True,
            plan_steps=["Lendo arquivo", f"{row_count} linhas, {len(headers)} colunas"],
            message=f"Planilha com {row_count} linhas. Colunas: {', '.join(str(h) for h in headers)}",
            data=SheetInfo(
                file=file_path,
                sheets=["Sheet1"],
                columns=[str(h) for h in headers],
                row_count=row_count,
            ),
        )
    except ImportError:
        return SpreadsheetResult(
            success=False,
            plan_steps=["Verificando dependências"],
            message="Suporte a .ods requer pandas + odfpy. Use .xlsx ou .csv.",
            error="missing_dependency",
        )


def _search_ods(file_path: str, query: str) -> list[CellResult]:
    """Search ODS file (basic support via pandas)."""
    try:
        import pandas as pd

        df = pd.read_excel(file_path, engine="odf")
        results: list[CellResult] = []

        for _col_idx, col in enumerate(df.columns):
            for row_idx, val in enumerate(df[col], start=2):
                if val is not None and query in str(val).lower():
                    results.append(
                        CellResult(
                            file=Path(file_path).name,
                            sheet="Sheet1",
                            row=row_idx,
                            column=str(col),
                            value=str(val),
                        )
                    )

        return results
    except ImportError:
        return []
